import streamlit as st
import pandas as pd
import io
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Meesho Reconciliation Tool", page_icon="🧾", layout="wide")

st.markdown("""
<style>
    .main-title { font-size: 2rem; font-weight: 700; color: #6C3FC5; margin-bottom: 0; }
    .sub-title  { font-size: 1rem; color: #888; margin-bottom: 1.5rem; }
    .section-header { font-size: 1.1rem; font-weight: 600; color: #444;
                      border-left: 4px solid #6C3FC5; padding-left: 10px;
                      margin: 1.2rem 0 0.6rem 0; }
    .stat-box   { background: #f8f5ff; border-radius: 10px; padding: 14px 18px;
                  border: 1px solid #e0d4f7; text-align: center; }
    .stat-num   { font-size: 1.6rem; font-weight: 700; color: #6C3FC5; }
    .stat-label { font-size: 0.78rem; color: #777; margin-top: 2px; }
    .profit     { color: #1a8c4e; font-weight: 600; }
    .loss       { color: #d93025; font-weight: 600; }
    .price-found { padding: 6px 12px; background: #e6f9ee; border-radius: 6px;
                   color: #1a8c4e; font-weight: 700; font-size: 1rem; }
    .price-not-found { padding: 6px 12px; background: #fff0f0; border-radius: 6px;
                       color: #d93025; font-weight: 600; font-size: 0.9rem; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">🧾 Meesho Reconciliation Tool</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Automated price reconciliation · YG · PE · AG accounts</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def clean_col(c):
    return str(c).strip().strip("'\"").strip()

def find_col(df, *keywords):
    cleaned = {clean_col(c).upper(): c for c in df.columns}
    for kw in keywords:
        match = next((cleaned[k] for k in cleaned if kw.upper() in k), None)
        if match:
            return match
    return None

def clean_df_columns(df):
    df.columns = [clean_col(c) for c in df.columns]
    return df

def detect_account(filename):
    fn = filename.upper()
    if '_YG' in fn or fn.endswith('YG.CSV'):
        return 'Yash Gallery', 'YG'
    elif '_PE' in fn or fn.endswith('PE.CSV'):
        return 'Pushpa', 'PE'
    elif '_AG' in fn or fn.endswith('AG.CSV'):
        return 'Ashirwad Garments', 'AG'
    return 'Unknown', 'UNK'

SIZE_NORMALIZE = {
    'FREE-SIZE: 36-40': 'F', 'FREE-SIZE: 28-32': 'F',
    'FREE-SIZE: 32-36': 'F', 'FREE-SIZE: 40-44': 'F',
    'FREE SIZE': 'F', 'FREE': 'F', 'FREESIZE': 'F',
    'XXXL': '3XL', 'XXXXL': '4XL', 'XXXXXL': '5XL',
}

def normalize_size(size):
    return SIZE_NORMALIZE.get(size.strip().upper(), size.strip())

SIZE_RANGE_MAP = {
    'S':   ['S-M', 'XS-S'],
    'M':   ['S-M', 'M-L'],
    'L':   ['L-XL', 'M-L'],
    'XL':  ['L-XL', 'XL-XXL'],
    'XXL': ['XXL-3XL', 'XL-XXL'],
    '3XL': ['XXL-3XL', '3XL-4XL'],
    '4XL': ['4XL-5XL', '3XL-4XL'],
    '5XL': ['4XL-5XL', '5XL-6XL'],
    '6XL': ['6XL-7XL', '5XL-6XL'],
    '7XL': ['6XL-7XL', '7XL-8XL'],
    '8XL': ['7XL-8XL'],
    'F':   ['F'],
}

PREFIX_REPLACEMENTS = [
    (r'YKN',     'YK'),
    (r'YKO251',  'YK251'),
    (r'YPLK',    'YK'),
    (r'PLYK',    'YK'),
    (r'YKC',     'YK'),      # NEW: YKC → YK transformation
    (r'YK-(\d)', r'YK\1'),
]

# ─────────────────────────────────────────────
# LOAD REFERENCE FILES
# ─────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_replace_sku(file_bytes):
    xl     = pd.ExcelFile(io.BytesIO(file_bytes))
    result = {}
    sheet_config = {
        'Meesho YG':     ('YG',  ['SELLER', 'MEESHO', 'SKU'],  ['OMS']),
        'Meesho Pushpa': ('PE',  ['MESSHO', 'MEESHO', 'SKU'],  ['OMS']),
        'Messho Ag':     ('AG',  ['MESSHO', 'MEESHO', 'SKU'],  ['OMS']),
    }
    for sheet, (acct, src_kws, oms_kws) in sheet_config.items():
        if sheet not in xl.sheet_names:
            continue
        df = xl.parse(sheet)
        df = clean_df_columns(df)
        src_col = find_col(df, *src_kws)
        oms_col = find_col(df, *oms_kws)
        if src_col is None or oms_col is None:
            st.warning(f"⚠️ Could not find columns in sheet [{sheet}]. Found: {df.columns.tolist()}")
            result[acct] = {}
            continue
        df = df[[src_col, oms_col]].dropna(subset=[src_col, oms_col])
        df[src_col] = df[src_col].astype(str).str.strip()
        df[oms_col] = df[oms_col].astype(str).str.strip()
        df = df[df[src_col] != 'nan']
        df = df[df[oms_col] != 'nan']
        result[acct] = dict(zip(df[src_col], df[oms_col]))
    return result

@st.cache_data(show_spinner=False)
def load_pwn(file_bytes):
    df = pd.read_excel(io.BytesIO(file_bytes), header=2)
    df = clean_df_columns(df)
    child_col = find_col(df, 'CHILD')
    pwn_col   = find_col(df, 'PWN')
    if child_col is None: child_col = df.columns[1]
    if pwn_col is None:   pwn_col   = df.columns[2]
    df = df[[child_col, pwn_col]].dropna(subset=[child_col, pwn_col])
    df[child_col] = df[child_col].astype(str).str.strip()
    df[pwn_col]   = pd.to_numeric(df[pwn_col], errors='coerce')
    df = df.dropna(subset=[pwn_col])
    df = df[df[child_col] != 'nan']
    exact_map = df.set_index(child_col)[pwn_col].to_dict()
    ci_map    = {k.lower(): (k, v) for k, v in exact_map.items()}
    return exact_map, ci_map

@st.cache_data(show_spinner=False)
def load_closed_sku(file_bytes):
    xl     = pd.ExcelFile(io.BytesIO(file_bytes))
    closed = {}
    def add_prices(sheet_name, sku_col_idx, price_col_idx):
        if sheet_name in xl.sheet_names:
            df = xl.parse(sheet_name, usecols=[sku_col_idx, price_col_idx])
            df = clean_df_columns(df)
            df.columns = ['SKU', 'Price']
            df = df.dropna(subset=['SKU', 'Price'])
            df['SKU']   = df['SKU'].astype(str).str.strip()
            df['Price'] = pd.to_numeric(df['Price'], errors='coerce')
            df = df.dropna(subset=['Price'])
            df = df[df['SKU'] != 'nan']
            for _, row in df.iterrows():
                closed.setdefault(str(row['SKU']), []).append(row['Price'])
    add_prices('Sheet1', 0, 1)
    add_prices('Sheet2', 0, 1)
    return {sku: min(prices) for sku, prices in closed.items() if prices}

# ─────────────────────────────────────────────
# PWN LOOKUP WITH ALL FALLBACKS
# ─────────────────────────────────────────────
def lookup_pwn(oms_sku, exact_map, ci_map):
    def try_key(key, note):
        if key in exact_map:
            return exact_map[key], key, note
        if key.lower() in ci_map:
            orig_key, price = ci_map[key.lower()]
            return price, orig_key, note + ' (ci)'
        return None, None, None

    p, k, n = try_key(oms_sku, 'Direct')
    if p is not None: return p, k, n

    for pattern, repl in PREFIX_REPLACEMENTS:
        candidate = re.sub(pattern, repl, oms_sku, count=1)
        if candidate != oms_sku:
            p, k, n = try_key(candidate, f'Prefix: {pattern}→{repl}')
            if p is not None: return p, k, n

    m = re.match(r'^(.+)-([A-Z0-9]+)-([A-Z0-9]+)$', oms_sku)
    if m:
        base, s1, s2 = m.group(1), m.group(2), m.group(3)
        for sz in [s1, s2]:
            p, k, n = try_key(f'{base}-{sz}', f'Split→{sz}')
            if p is not None: return p, k, n

    m2 = re.match(r'^(.+)-([A-Z0-9]+)$', oms_sku)
    if m2:
        base, size = m2.group(1), m2.group(2)
        for range_sz in SIZE_RANGE_MAP.get(size, []):
            p, k, n = try_key(f'{base}-{range_sz}', f'Range: {size}→{range_sz}')
            if p is not None: return p, k, n

    return None, oms_sku, 'Not Found'

# ─────────────────────────────────────────────
# PROCESS ONE CSV
# ─────────────────────────────────────────────
def process_csv(df_raw, account_code, company_name, sku_map, exact_map, ci_map, closed_map):
    acct_sku_map = sku_map.get(account_code, {})
    records      = []

    for _, row in df_raw.iterrows():
        reason       = row.get('Reason for Credit Entry', '')
        sub_order    = row.get('Sub Order No', '')
        order_date   = row.get('Order Date', '')
        state        = row.get('Customer State', '')
        product_name = row.get('Product Name', '')
        raw_sku      = str(row.get('SKU', '')).strip()
        size         = str(row.get('Size', '')).strip()
        qty          = pd.to_numeric(row.get('Quantity', 1), errors='coerce')
        qty          = int(qty) if pd.notna(qty) else 1
        listed_price = pd.to_numeric(row.get('Supplier Listed Price (Incl. GST + Commission)', 0), errors='coerce') or 0
        disc_price   = pd.to_numeric(row.get('Supplier Discounted Price (Incl GST and Commision)', 0), errors='coerce') or 0
        packet_id    = row.get('Packet Id', '')

        size_clean      = size if size and size.lower() not in ('nan', '', 'none') else ''
        meesho_sku      = f'{raw_sku}-{size_clean}' if size_clean else raw_sku
        size_norm       = normalize_size(size_clean) if size_clean else ''
        meesho_sku_norm = f'{raw_sku}-{size_norm}' if size_norm else raw_sku

        if meesho_sku in acct_sku_map:
            oms_sku = acct_sku_map[meesho_sku]
        elif meesho_sku_norm in acct_sku_map:
            oms_sku = acct_sku_map[meesho_sku_norm]
        else:
            oms_sku = meesho_sku_norm

        sku_status   = 'On Going'
        pwn_10_price = None
        final_oms    = oms_sku
        lookup_note  = ''

        if oms_sku in closed_map:
            sku_status   = 'Closed'
            pwn_10_price = closed_map[oms_sku]
            lookup_note  = 'Closed SKU'
        else:
            pwn_val, final_oms, lookup_note = lookup_pwn(oms_sku, exact_map, ci_map)
            pwn_10_price = pwn_val if pwn_val is not None else 'SKU Not Found'

        disc_price_qty = round(disc_price * qty, 2)
        if isinstance(pwn_10_price, (int, float)):
            pwn_10_qty = round(pwn_10_price * qty, 2)
            difference = round(disc_price_qty - pwn_10_qty, 2)
        else:
            pwn_10_qty = 'SKU Not Found'
            difference = 'SKU Not Found'

        records.append({
            'Company':                         company_name,
            'Reason for Credit Entry':         reason,
            'Sub Order No':                    sub_order,
            'Order Date':                      order_date,
            'Customer State':                  state,
            'Product Name':                    product_name,
            'Meesho SKU':                      meesho_sku,
            'OMS SKU':                         oms_sku,
            'Final OMS SKU':                   final_oms,
            'Size':                            size,
            'Quantity':                        qty,
            'Supplier Listed Price':           listed_price,
            'Supplier Discounted Price':       disc_price,
            'Supplier Discounted Price * Qty': disc_price_qty,
            'Update PWN+10%':                  pwn_10_price,
            'Update PWN+10% * Qty':            pwn_10_qty,
            'Difference Amount':               difference,
            'SKU Status':                      sku_status,
            'Lookup Note':                     lookup_note,
            'Packet Id':                       packet_id,
        })

    return pd.DataFrame(records)

# ─────────────────────────────────────────────
# APPLY CORRECTIONS TO A DATAFRAME
# ─────────────────────────────────────────────
def apply_corrections_to_df(df, corrections):
    """Apply saved SKU corrections to a dataframe copy and return it."""
    df = df.copy()
    for orig_sku, fix in corrections.items():
        # Match rows that are still SKU Not Found AND have this original Final OMS SKU
        row_mask = (df['Difference Amount'] == 'SKU Not Found') & (df['Final OMS SKU'] == orig_sku)
        if row_mask.any():
            qty_vals  = df.loc[row_mask, 'Quantity']
            disc_vals = df.loc[row_mask, 'Supplier Discounted Price']
            df.loc[row_mask, 'Final OMS SKU']                    = fix['corrected_sku']
            df.loc[row_mask, 'Update PWN+10%']                   = fix['price']
            df.loc[row_mask, 'Update PWN+10% * Qty']             = (fix['price'] * qty_vals).round(2)
            df.loc[row_mask, 'Supplier Discounted Price * Qty']  = (disc_vals * qty_vals).round(2)
            df.loc[row_mask, 'Difference Amount']                = (disc_vals * qty_vals - fix['price'] * qty_vals).round(2)
            df.loc[row_mask, 'Lookup Note']                      = fix['note'] + ' (corrected)'
            df.loc[row_mask, 'SKU Status']                       = 'On Going'
    return df

# ─────────────────────────────────────────────
# EXPORT STYLED EXCEL
# ─────────────────────────────────────────────
def export_excel(sheets_dict):
    wb = Workbook()
    wb.remove(wb.active)

    HEADER_FILL   = PatternFill('solid', fgColor='6C3FC5')
    PROFIT_FILL   = PatternFill('solid', fgColor='C6EFCE')
    LOSS_FILL     = PatternFill('solid', fgColor='FFC7CE')
    NOTFOUND_FILL = PatternFill('solid', fgColor='FFEB9C')
    CLOSED_FILL   = PatternFill('solid', fgColor='BDD7EE')
    ALT_FILL      = PatternFill('solid', fgColor='F3EFFF')
    WHITE_FILL    = PatternFill('solid', fgColor='FFFFFF')
    HEADER_FONT   = Font(bold=True, color='FFFFFF', size=10)
    NORMAL_FONT   = Font(size=10)
    thin          = Side(style='thin', color='CCCCCC')
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    MONEY_COLS = {
        'Supplier Listed Price', 'Supplier Discounted Price',
        'Supplier Discounted Price * Qty', 'Update PWN+10%',
        'Update PWN+10% * Qty', 'Difference Amount'
    }

    for sheet_name, df in sheets_dict.items():
        ws   = wb.create_sheet(title=sheet_name[:31])
        cols = list(df.columns)

        for c_idx, col in enumerate(cols, 1):
            cell           = ws.cell(1, c_idx, col)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border    = border
        ws.row_dimensions[1].height = 38

        diff_col_idx = cols.index('Difference Amount') + 1 if 'Difference Amount' in cols else None

        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            base_fill  = ALT_FILL if r_idx % 2 == 0 else WHITE_FILL
            diff_val   = row.get('Difference Amount', None)
            status_val = str(row.get('SKU Status', ''))

            for c_idx, col in enumerate(cols, 1):
                val  = row[col]
                cell = ws.cell(r_idx, c_idx, val)

                if val == 'SKU Not Found':
                    cell.fill = NOTFOUND_FILL
                elif status_val == 'Closed':
                    cell.fill = CLOSED_FILL
                elif c_idx == diff_col_idx and isinstance(diff_val, (int, float)):
                    cell.fill = PROFIT_FILL if diff_val >= 0 else LOSS_FILL
                else:
                    cell.fill = base_fill

                cell.font   = NORMAL_FONT
                cell.border = border
                if col in MONEY_COLS and isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(vertical='center')

        for c_idx, col in enumerate(cols, 1):
            max_len = len(str(col))
            for r_idx in range(2, ws.max_row + 1):
                v = ws.cell(r_idx, c_idx).value
                if v:
                    max_len = max(max_len, min(len(str(v)), 50))
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 45)

        ws.freeze_panes    = 'A2'
        ws.auto_filter.ref = ws.dimensions

    ws_s        = wb.create_sheet(title='Summary', index=0)
    sum_headers = ['Account', 'Company', 'Total Orders', 'Profit Orders',
                   'Loss Orders', 'SKU Not Found', 'Closed SKUs', 'Total Difference (₹)']
    for c, h in enumerate(sum_headers, 1):
        cell           = ws_s.cell(1, c, h)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border    = border
    ws_s.row_dimensions[1].height = 30

    for r, (sheet_name, df) in enumerate(sheets_dict.items(), 2):
        diff_series = pd.to_numeric(df['Difference Amount'], errors='coerce')
        company     = df['Company'].iloc[0] if len(df) > 0 else ''
        acct_code   = sheet_name.split('_')[0]
        not_found   = int((df['Difference Amount'] == 'SKU Not Found').sum())
        closed_cnt  = int((df['SKU Status'] == 'Closed').sum())
        profit_cnt  = int((diff_series >= 0).sum())
        loss_cnt    = int((diff_series < 0).sum())
        total_diff  = round(diff_series.sum(), 2)

        for c, val in enumerate([acct_code, company, len(df), profit_cnt,
                                  loss_cnt, not_found, closed_cnt, total_diff], 1):
            cell           = ws_s.cell(r, c, val)
            cell.border    = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if c == 8 and isinstance(val, float):
                cell.number_format = '#,##0.00'
                cell.font = Font(bold=True, color='1a8c4e' if val >= 0 else 'd93025')

    for c in range(1, len(sum_headers) + 1):
        ws_s.column_dimensions[get_column_letter(c)].width = 22
    ws_s.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown('### 📁 Reference Files')
    st.caption('Upload once — reused for all reconciliations')
    ref_replace = st.file_uploader('Replace SKU (.xlsx)',       type='xlsx', key='ref1')
    ref_pwn     = st.file_uploader('PWN + 10% (.xlsx)',         type='xlsx', key='ref2')
    ref_closed  = st.file_uploader('Meesho Closed SKU (.xlsx)', type='xlsx', key='ref3')
    st.markdown('---')
    st.markdown('### ℹ️ Account Detection')
    st.markdown("""
    Filename must contain:
    - `_YG` → **Yash Gallery**
    - `_PE` → **Pushpa**
    - `_AG` → **Ashirwad Garments**
    """)
    st.markdown('---')
    st.markdown('### 🎨 Excel Colour Guide')
    st.markdown("""
    - 🟢 Green = Profit
    - 🔴 Red = Loss
    - 🔵 Blue = Closed SKU
    - 🟡 Yellow = SKU Not Found
    """)
    st.markdown('---')
    st.markdown('### 🔄 SKU Resolution Steps')
    st.markdown("""
    1. Build: `SKU + - + Size` = Meesho SKU
    2. Normalize: `Free-Size:36-40`→`F`, `XXXL`→`3XL`
    3. Replace map (YG/PE/AG) → OMS SKU
    4. Direct PWN lookup + case-insensitive
    5. Prefix fix: `PLYK/YKN/YPLK`→`YK`
    6. Split combined size: `3XL-4XL`→`3XL`
    7. Size range: `L`→`L-XL`, `XXL`→`XXL-3XL`
    """)

# ─────────────────────────────────────────────
# SESSION STATE INIT
# ─────────────────────────────────────────────
if 'reconciliation_done' not in st.session_state:
    st.session_state.reconciliation_done = False
if 'raw_sheets' not in st.session_state:
    st.session_state.raw_sheets = {}       # sheet_key → original df_out (never modified)
if 'corrections' not in st.session_state:
    st.session_state.corrections = {}      # sheet_key → {orig_sku: {corrected_sku, price, note}}
if 'ref_maps' not in st.session_state:
    st.session_state.ref_maps = {}         # exact_map, ci_map, closed_map
if 'file_meta' not in st.session_state:
    st.session_state.file_meta = {}        # sheet_key → {company_name, account_code, filename}

# ─────────────────────────────────────────────
# MAIN — UPLOAD & RUN
# ─────────────────────────────────────────────
st.markdown('<div class="section-header">📂 Upload Meesho Order CSV Files</div>', unsafe_allow_html=True)
order_files = st.file_uploader(
    'Upload one or more order CSV files',
    type='csv', accept_multiple_files=True, key='orders'
)

if st.button('🚀 Run Reconciliation', type='primary', use_container_width=True):
    errors = []
    if not ref_replace: errors.append('Replace SKU file')
    if not ref_pwn:     errors.append('PWN+10% file')
    if not ref_closed:  errors.append('Meesho Closed SKU file')
    if not order_files: errors.append('At least one order CSV')

    if errors:
        st.error(f"Please upload: {', '.join(errors)}")
    else:
        replace_bytes = ref_replace.read()
        pwn_bytes     = ref_pwn.read()
        closed_bytes  = ref_closed.read()

        with st.spinner('Loading reference files...'):
            sku_map           = load_replace_sku(replace_bytes)
            exact_map, ci_map = load_pwn(pwn_bytes)
            closed_map        = load_closed_sku(closed_bytes)

        # Store reference maps in session state so correction UI can use them
        st.session_state.ref_maps = {
            'exact_map': exact_map,
            'ci_map':    ci_map,
            'closed_map': closed_map,
        }

        yg_size = len(sku_map.get('YG', {}))
        pe_size = len(sku_map.get('PE', {}))
        ag_size = len(sku_map.get('AG', {}))

        st.success(
            f"✅ Loaded — **PWN: {len(exact_map):,} SKUs** | "
            f"**Closed: {len(closed_map):,} SKUs** | "
            f"Replace → YG: {yg_size:,} | PE: {pe_size:,} | AG: {ag_size:,}"
        )
        if yg_size == 0 or pe_size == 0 or ag_size == 0:
            st.warning('⚠️ One or more Replace SKU maps loaded empty! Check the Replace_SKU.xlsx file.')

        # Reset previous results
        st.session_state.raw_sheets  = {}
        st.session_state.corrections = {}
        st.session_state.file_meta   = {}

        progress = st.progress(0)
        for i, f in enumerate(order_files):
            company_name, account_code = detect_account(f.name)
            try:
                df_raw    = pd.read_csv(io.BytesIO(f.read()))
                df_out    = process_csv(df_raw, account_code, company_name,
                                        sku_map, exact_map, ci_map, closed_map)
                sheet_key = f'{account_code}_{f.name[:25]}'
                # Store original result — never mutate this
                st.session_state.raw_sheets[sheet_key]  = df_out
                st.session_state.corrections[sheet_key] = {}
                st.session_state.file_meta[sheet_key]   = {
                    'company_name':  company_name,
                    'account_code':  account_code,
                    'filename':      f.name,
                    'total_input':   len(df_raw),
                }
            except Exception as e:
                import traceback
                st.error(f'❌ Error processing {f.name}: {e}')
                st.code(traceback.format_exc())
            progress.progress((i + 1) / len(order_files))

        st.session_state.reconciliation_done = True
        st.rerun()

# ─────────────────────────────────────────────
# RESULTS SECTION — shown after reconciliation
# ─────────────────────────────────────────────
if st.session_state.reconciliation_done and st.session_state.raw_sheets:

    exact_map  = st.session_state.ref_maps.get('exact_map', {})
    ci_map     = st.session_state.ref_maps.get('ci_map', {})
    closed_map = st.session_state.ref_maps.get('closed_map', {})

    all_stats      = []
    corrected_sheets = {}   # final sheets with corrections applied — used for Excel export

    for sheet_key, df_orig in st.session_state.raw_sheets.items():
        meta         = st.session_state.file_meta[sheet_key]
        company_name = meta['company_name']
        account_code = meta['account_code']
        filename     = meta['filename']
        total_input  = meta['total_input']
        corrections  = st.session_state.corrections.get(sheet_key, {})

        # Apply any saved corrections to get the working df
        df_out = apply_corrections_to_df(df_orig, corrections)
        corrected_sheets[sheet_key] = df_out

        st.markdown(
            f'<div class="section-header">📄 {filename} → {company_name} ({account_code})</div>',
            unsafe_allow_html=True
        )

        total_output = len(df_out)
        diff_series  = pd.to_numeric(df_out['Difference Amount'], errors='coerce')
        total_diff   = round(diff_series.sum(), 2)
        profit_cnt   = int((diff_series >= 0).sum())
        loss_cnt     = int((diff_series < 0).sum())
        not_found    = int((df_out['Difference Amount'] == 'SKU Not Found').sum())
        closed_cnt   = int((df_out['SKU Status'] == 'Closed').sum())

        all_stats.append({
            'account': account_code, 'company': company_name,
            'orders': total_output,  'profit': profit_cnt,
            'loss': loss_cnt,        'not_found': not_found,
            'closed': closed_cnt,    'total_diff': total_diff
        })

        if total_input != total_output:
            st.warning(f'⚠️ Input: {total_input} rows | Output: {total_output} rows')
        else:
            st.success(f'✅ All {total_input:,} rows processed — no rows dropped')

        c1, c2, c3, c4, c5, c6 = st.columns(6)
        for col, num, label, cls in [
            (c1, f'{total_output:,}',  'Total Orders',   ''),
            (c2, f'{profit_cnt:,}',    'Profit Orders',  'profit'),
            (c3, f'{loss_cnt:,}',      'Loss Orders',    'loss'),
            (c4, f'{closed_cnt:,}',    'Closed SKUs',    ''),
            (c5, f'{not_found:,}',     'SKU Not Found',  'loss' if not_found else ''),
            (c6, f'₹{total_diff:,.0f}','Net Difference', 'profit' if total_diff >= 0 else 'loss'),
        ]:
            with col:
                st.markdown(
                    f'<div class="stat-box"><div class="stat-num {cls}">{num}</div>'
                    f'<div class="stat-label">{label}</div></div>',
                    unsafe_allow_html=True
                )

        # ── SKU NOT FOUND — inspect table ───────────────────────
        # Use the ORIGINAL df to find which SKUs still need correcting
        nf_orig = df_orig[df_orig['Difference Amount'] == 'SKU Not Found'][
            ['Meesho SKU', 'OMS SKU', 'Final OMS SKU', 'Size', 'Lookup Note']
        ].drop_duplicates()

        if len(nf_orig) > 0:
            with st.expander(
                f'⚠️ {not_found} SKU Not Found — click to inspect'
                + (f' ({len(corrections)} corrected ✅)' if corrections else ''),
                expanded=(not_found > 0)
            ):
                st.dataframe(nf_orig, use_container_width=True)

            # ── SKU CORRECTION UI ────────────────────────────────
            st.markdown(
                '<div class="section-header">✏️ Correct SKU Not Found</div>',
                unsafe_allow_html=True
            )
            st.caption(
                'Type the correct Final OMS SKU — price is looked up **instantly**. '
                'Click **✅ Save** to apply. Changes reflect in the Excel download immediately.'
            )

            unique_not_found = nf_orig['Final OMS SKU'].unique().tolist()

            for orig_sku in unique_not_found:
                already_fixed = orig_sku in corrections
                saved_fix     = corrections.get(orig_sku, {})

                col_a, col_b, col_c, col_d = st.columns([3, 3, 2, 1])

                with col_a:
                    # Show with green tick if already corrected
                    label = f'{"✅ " if already_fixed else ""}Original SKU'
                    st.text_input(label, value=orig_sku, disabled=True,
                                  key=f'orig__{sheet_key}__{orig_sku}')

                with col_b:
                    input_key = f'fix__{sheet_key}__{orig_sku}'
                    corrected_input = st.text_input(
                        'Correct Final OMS SKU',
                        value=saved_fix.get('corrected_sku', ''),
                        key=input_key,
                        placeholder='e.g. YK251-XL'
                    )

                # Live lookup every render based on what's typed right now
                typed_sku       = corrected_input.strip()
                looked_up_price = None
                looked_up_note  = ''

                if typed_sku:
                    looked_up_price, _, looked_up_note = lookup_pwn(typed_sku, exact_map, ci_map)
                    if looked_up_price is None and typed_sku in closed_map:
                        looked_up_price = closed_map[typed_sku]
                        looked_up_note  = 'Closed SKU'

                with col_c:
                    st.write('')  # spacing
                    if typed_sku:
                        if looked_up_price is not None:
                            st.markdown(
                                f'<div class="price-found">₹{looked_up_price:,.2f}'
                                f'<span style="font-size:0.72rem;color:#555;margin-left:8px">'
                                f'{looked_up_note}</span></div>',
                                unsafe_allow_html=True
                            )
                        else:
                            st.markdown(
                                '<div class="price-not-found">❌ SKU not in PWN file</div>',
                                unsafe_allow_html=True
                            )
                    elif already_fixed:
                        st.markdown(
                            f'<div class="price-found">₹{saved_fix["price"]:,.2f} ✅</div>',
                            unsafe_allow_html=True
                        )

                with col_d:
                    st.write('')  # spacing
                    save_key = f'save__{sheet_key}__{orig_sku}'
                    # Button is enabled only when a valid price is found
                    if st.button('✅ Save', key=save_key,
                                 disabled=(looked_up_price is None or not typed_sku)):
                        st.session_state.corrections[sheet_key][orig_sku] = {
                            'corrected_sku': typed_sku,
                            'price':         looked_up_price,
                            'note':          looked_up_note,
                        }
                        st.rerun()

            # Show summary of applied corrections
            if corrections:
                st.success(
                    f'✅ {len(corrections)} of {len(unique_not_found)} SKUs corrected — '
                    f'reflected in preview and Excel download below.'
                )
                if st.button(f'🔄 Clear all corrections for {account_code}',
                             key=f'clear_{sheet_key}'):
                    st.session_state.corrections[sheet_key] = {}
                    st.rerun()

        with st.expander(f'👁️ Preview — {company_name} (first 20 rows, with corrections applied)'):
            st.dataframe(df_out.head(20), use_container_width=True)

        st.markdown('---')

    # ── OVERALL SUMMARY ──────────────────────────────────────────
    if all_stats:
        st.markdown('<div class="section-header">📊 Overall Summary</div>', unsafe_allow_html=True)
        grand_orders = sum(s['orders'] for s in all_stats)
        grand_profit = sum(s['profit'] for s in all_stats)
        grand_loss   = sum(s['loss']   for s in all_stats)
        grand_diff   = round(sum(s['total_diff'] for s in all_stats), 2)

        c1, c2, c3, c4 = st.columns(4)
        for col, num, label, cls in [
            (c1, f'{grand_orders:,}',  'Total Orders (All)',  ''),
            (c2, f'{grand_profit:,}',  'Total Profit Orders', 'profit'),
            (c3, f'{grand_loss:,}',    'Total Loss Orders',   'loss'),
            (c4, f'₹{grand_diff:,.0f}','Grand Net Difference','profit' if grand_diff >= 0 else 'loss'),
        ]:
            with col:
                st.markdown(
                    f'<div class="stat-box"><div class="stat-num {cls}">{num}</div>'
                    f'<div class="stat-label">{label}</div></div>',
                    unsafe_allow_html=True
                )

    # ── DOWNLOAD — uses corrected data ───────────────────────────
    st.markdown('---')
    excel_buf = export_excel(corrected_sheets)
    st.download_button(
        label='📥 Download Excel Reconciliation Report',
        data=excel_buf,
        file_name='Meesho_Reconciliation_Report.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        use_container_width=True,
        type='primary'
    )

st.markdown('---')
st.markdown(
    "<center style='color:#aaa;font-size:0.8rem'>Meesho Reconciliation Tool · Built with Streamlit</center>",
    unsafe_allow_html=True
)
